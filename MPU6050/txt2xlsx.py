import re
import openpyxl

# 创建新表格
print('创建表格...')
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = '角度'
ws['B1'] = '角速度'
ws['C1'] = '角加速度'
print('创建表格完毕！')
print('读取数据...')
f = open('步态数据.txt', 'r') # 读取得到的txt文件
print('读取数据完毕！')
lines = f.readlines()
count = 2 # 表格待写行数
print('写入数据到表格中...')
for line in lines:
    # 将每一行的数据分割为多个字符串
    items = re.sub('\n|\*', '', line)
    items = re.split(':' ,items)
    # print(items)
    # 分别将每一行的三个数据录入表格
    if items is not ['']:
        try:
            ws.cell(row=count, column=1).value = float(items[0]) # 角度
            ws.cell(row=count, column=2).value = float(items[1]) # 角速度
            ws.cell(row=count, column=3).value = float(items[2]) # 角加速度
            count = count + 1
        except:
            pass
print('写入数据完毕！')
print('保存表格...')
wb.save('步态数据.xlsx')
print('Done!')